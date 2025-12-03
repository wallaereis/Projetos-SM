import pyautogui as py
import time
import openpyxl as xl

py.FAILSAFE = True
py.PAUSE = 1



df = xl.load_workbook(r"C:\Users\wallace.souza\Downloads\teste.xlsx", data_only=True)

ws = df["Plan1"]

print("o programa começou")

py.hotkey('alt' , 'tab')



for row in ws.iter_rows(min_row=2):
    py.press("f2")
    py.write(str(row[0].value))
    py.press("f8")
    # Empresa/Segmento Abaixo
    py.click(627,184)

#     # # CD ABAIXO
#     # py.click(184,602)
#     # # BOTAO
#     # py.click(218,600)
#     # # ATIVAR
#     # py.click(200,607)


#     # CAFUBA ABAIXO
#     py.click(194,688)
#     # BOTAO
#     py.click(224,686) 
#     # ATIVO
#     py.click(216,697) 

    # CAMPO GRANDE ABAIXO
    py.click(196,703)
    # BOTAO 
    py.click(220,702)
    # ATIVO
    py.click(194,713)

    py.press("f4")



#     # for x in range(3):
#     #     py.write("a")
#     # py.press("f4")     



# for row in ws.iter_rows(min_row=2):
#     py.press("f2")
#     py.write(str(row[0].value))
#     py.press("f8")
#     # FAMILIA
#     py.click(619,223)
#     # MIX
#     py.click(255,134)
#     # COMPENSACAO
#     py.click(275,197)
#     # EXCLUI
#     py.click(382,97)
#     # SALVA
#     py.press("f4")
#     # ADICIONA
#     py.click(354,97)
#     # INSERE
#     py.doubleClick(215,164)
#     # SALVA
#     py.press("f4")
#     # APERTAR PRA SAIR
#     py.click(1207,62)

    # py.press("f4")

 
print("o programa terminou")


